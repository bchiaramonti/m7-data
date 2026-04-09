#!/usr/bin/env python3
"""
Gera Tabela de Desconexoes em Excel (.xlsx) com branding M7-2026.

Uso:
    python generate-disconnection-table.py \
        --input data.json \
        --output processo-desconexoes.xlsx \
        --logo assets/m7-logo-dark.png

Input JSON esperado:
{
    "metadata": {
        "processName": "...",
        "code": "...",
        "responsible": "...",
        "level": "...",
        "date": "...",
        "version": "..."
    },
    "interfaces": [
        { "id": "I1", "zone": "input", "artifact": "...", "provider": "...",
          "receiver": "...", "status": "conforme|melhoria|neutral", "note": "..." }
    ],
    "disconnections": [
        { "id": "I1", "zone": "input", "item": "...", "providerReceiver": "...",
          "disconnection": "...", "impact": "Alto|Medio|Baixo", "suggestedAction": "..." }
    ]
}
"""

import argparse
import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl nao encontrado. Instale com: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# --- M7-2026 Design Tokens ---
VERDE_CAQUI = "424135"
OFF_WHITE = "fffdef"
LIME = "EEF77C"
SUCCESS = "4CAF50"
ERROR = "E46962"
WHITE = "FFFFFF"
NEUTRAL_GRAY = "8a8981"

# Fonts (Arial como fallback — openpyxl nao suporta fontes custom)
FONT_TITLE = Font(name="Arial", size=16, bold=True, color=VERDE_CAQUI)
FONT_SUBTITLE = Font(name="Arial", size=11, color=VERDE_CAQUI)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color=OFF_WHITE)
FONT_BODY = Font(name="Arial", size=10, color=VERDE_CAQUI)
FONT_BODY_WHITE = Font(name="Arial", size=10, bold=True, color=WHITE)
FONT_BODY_DARK = Font(name="Arial", size=10, bold=True, color=VERDE_CAQUI)
FONT_FOOTER = Font(name="Arial", size=8, italic=True, color=NEUTRAL_GRAY)

# Fills
FILL_HEADER = PatternFill(start_color=VERDE_CAQUI, end_color=VERDE_CAQUI, fill_type="solid")
FILL_OFF_WHITE = PatternFill(start_color=OFF_WHITE, end_color=OFF_WHITE, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_ALTO = PatternFill(start_color=ERROR, end_color=ERROR, fill_type="solid")
FILL_MEDIO = PatternFill(start_color=LIME, end_color=LIME, fill_type="solid")
FILL_BAIXO = PatternFill(start_color=SUCCESS, end_color=SUCCESS, fill_type="solid")
FILL_CONFORME = PatternFill(start_color=SUCCESS, end_color=SUCCESS, fill_type="solid")
FILL_MELHORIA = PatternFill(start_color=ERROR, end_color=ERROR, fill_type="solid")
FILL_NEUTRAL = PatternFill(start_color=NEUTRAL_GRAY, end_color=NEUTRAL_GRAY, fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style="hair", color=VERDE_CAQUI),
    right=Side(style="hair", color=VERDE_CAQUI),
    top=Side(style="hair", color=VERDE_CAQUI),
    bottom=Side(style="hair", color=VERDE_CAQUI),
)

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_header(ws, metadata, logo_path, start_row=1):
    """Escreve header com logo e metadados do processo."""

    # Logo
    if logo_path and os.path.isfile(logo_path):
        img = Image(logo_path)
        img.height = 40
        img.width = int(img.width * (40 / img.height)) if img.height > 0 else 100
        ws.add_image(img, f"A{start_row}")
    ws.row_dimensions[start_row].height = 45

    # Titulo
    title_row = start_row + 1
    ws.merge_cells(f"A{title_row}:H{title_row}")
    cell = ws.cell(row=title_row, column=1)
    cell.value = f"Tabela de Desconexoes — {metadata.get('processName', 'Processo')}"
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_LEFT
    ws.row_dimensions[title_row].height = 28

    # Metadados linha 1
    meta_row1 = start_row + 2
    ws.merge_cells(f"A{meta_row1}:H{meta_row1}")
    cell = ws.cell(row=meta_row1, column=1)
    cell.value = (
        f"Codigo: {metadata.get('code', '—')} | "
        f"Responsavel: {metadata.get('responsible', '—')} | "
        f"Nivel: {metadata.get('level', '—')}"
    )
    cell.font = FONT_SUBTITLE
    cell.alignment = ALIGN_LEFT

    # Metadados linha 2
    meta_row2 = start_row + 3
    ws.merge_cells(f"A{meta_row2}:H{meta_row2}")
    cell = ws.cell(row=meta_row2, column=1)
    cell.value = (
        f"Data do levantamento: {metadata.get('date', '—')} | "
        f"Versao: {metadata.get('version', '—')}"
    )
    cell.font = FONT_SUBTITLE
    cell.alignment = ALIGN_LEFT

    # Linha separadora
    return start_row + 5


def write_table_header(ws, row, columns):
    """Escreve header de tabela com estilo M7."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_name
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def apply_impact_style(cell, impact):
    """Aplica formatacao condicional por impacto."""
    impact_lower = (impact or "").lower().strip()
    if impact_lower == "alto":
        cell.fill = FILL_ALTO
        cell.font = FONT_BODY_WHITE
    elif impact_lower in ("medio", "médio"):
        cell.fill = FILL_MEDIO
        cell.font = FONT_BODY_DARK
    elif impact_lower == "baixo":
        cell.fill = FILL_BAIXO
        cell.font = FONT_BODY_WHITE


def apply_status_style(cell, status):
    """Aplica formatacao condicional por status."""
    status_lower = (status or "").lower().strip()
    if status_lower == "conforme":
        cell.fill = FILL_CONFORME
        cell.font = FONT_BODY_WHITE
    elif status_lower == "melhoria":
        cell.fill = FILL_MELHORIA
        cell.font = FONT_BODY_WHITE
    elif status_lower == "neutral":
        cell.fill = FILL_NEUTRAL
        cell.font = FONT_BODY_WHITE


def write_data_row(ws, row, values, is_alt):
    """Escreve uma linha de dados com formatacao alternada."""
    fill = FILL_OFF_WHITE if is_alt else FILL_WHITE
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = val
        cell.font = FONT_BODY
        cell.fill = fill
        cell.alignment = ALIGN_LEFT if col_idx > 1 else ALIGN_CENTER
        cell.border = THIN_BORDER


def auto_width(ws, columns, min_width=10, max_width=45):
    """Ajusta largura das colunas automaticamente."""
    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(columns[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_disconnections_sheet(wb, data, logo_path):
    """Cria aba 'Desconexoes' com dados de gaps."""
    ws = wb.active
    ws.title = "Desconexoes"
    ws.sheet_properties.tabColor = ERROR

    metadata = data.get("metadata", {})
    disconnections = data.get("disconnections", [])

    # Header
    data_start = write_header(ws, metadata, logo_path)

    # Colunas da tabela
    columns = [
        "ID", "Interface", "Zona", "Item",
        "Provedor/Receptor", "Desconexao", "Impacto", "Acao Sugerida"
    ]
    write_table_header(ws, data_start, columns)

    # Dados
    for idx, disc in enumerate(disconnections):
        row = data_start + 1 + idx
        values = [
            disc.get("id", ""),
            disc.get("id", ""),
            translate_zone(disc.get("zone", "")),
            disc.get("item", ""),
            disc.get("providerReceiver", ""),
            disc.get("disconnection", ""),
            disc.get("impact", ""),
            disc.get("suggestedAction", ""),
        ]
        write_data_row(ws, row, values, idx % 2 == 0)

        # Formatacao de impacto
        impact_cell = ws.cell(row=row, column=7)
        apply_impact_style(impact_cell, disc.get("impact", ""))

    # Auto-filtro
    if disconnections:
        last_row = data_start + len(disconnections)
        ws.auto_filter.ref = f"A{data_start}:H{last_row}"

    # Freeze panes
    ws.freeze_panes = f"A{data_start + 1}"

    # Auto width
    auto_width(ws, columns)

    # Footer
    footer_row = data_start + len(disconnections) + 2
    ws.merge_cells(f"A{footer_row}:H{footer_row}")
    cell = ws.cell(row=footer_row, column=1)
    cell.value = "DEIP — Diagrama de Escopo, Interface e Processos | Metodologia BPM CBOK | M7 Investimentos"
    cell.font = FONT_FOOTER
    cell.alignment = ALIGN_LEFT


def create_all_interfaces_sheet(wb, data, logo_path):
    """Cria aba 'Todas as Interfaces' com lista completa."""
    ws = wb.create_sheet("Todas as Interfaces")
    ws.sheet_properties.tabColor = VERDE_CAQUI

    metadata = data.get("metadata", {})
    interfaces = data.get("interfaces", [])

    # Header
    data_start = write_header(ws, metadata, logo_path)

    # Colunas
    columns = ["ID", "Zona", "Item", "Provedor/Receptor", "Status"]
    write_table_header(ws, data_start, columns)

    # Dados
    for idx, iface in enumerate(interfaces):
        row = data_start + 1 + idx
        provider_receiver = iface.get("provider", "") or iface.get("receiver", "")
        values = [
            iface.get("id", ""),
            translate_zone(iface.get("zone", "")),
            iface.get("artifact", ""),
            provider_receiver,
            translate_status(iface.get("status", "")),
        ]
        write_data_row(ws, row, values, idx % 2 == 0)

        # Formatacao de status
        status_cell = ws.cell(row=row, column=5)
        apply_status_style(status_cell, iface.get("status", ""))

    # Auto-filtro
    if interfaces:
        last_row = data_start + len(interfaces)
        ws.auto_filter.ref = f"A{data_start}:E{last_row}"

    # Freeze panes
    ws.freeze_panes = f"A{data_start + 1}"

    # Auto width
    auto_width(ws, columns)


def create_summary_sheet(wb, data, logo_path):
    """Cria aba 'Resumo' com estatisticas."""
    ws = wb.create_sheet("Resumo")
    ws.sheet_properties.tabColor = SUCCESS

    metadata = data.get("metadata", {})
    interfaces = data.get("interfaces", [])
    disconnections = data.get("disconnections", [])

    # Header
    data_start = write_header(ws, metadata, logo_path)

    # --- Tabela 1: Interfaces por Zona ---
    ws.cell(row=data_start, column=1, value="Interfaces por Zona").font = Font(
        name="Arial", size=12, bold=True, color=VERDE_CAQUI
    )
    data_start += 1

    zone_cols = ["Zona", "Total", "Conforme", "Melhoria", "Nao Avaliado"]
    write_table_header(ws, data_start, zone_cols)

    zones = {"input": "Entradas (I)", "output": "Saidas (O)", "support": "Suporte (S)", "regulation": "Regulacao (R)"}
    row = data_start + 1
    totals = {"total": 0, "conforme": 0, "melhoria": 0, "neutral": 0}

    for zone_key, zone_label in zones.items():
        zone_ifaces = [i for i in interfaces if i.get("zone") == zone_key]
        total = len(zone_ifaces)
        conforme = sum(1 for i in zone_ifaces if i.get("status") == "conforme")
        melhoria = sum(1 for i in zone_ifaces if i.get("status") == "melhoria")
        neutral = total - conforme - melhoria

        values = [zone_label, total, conforme, melhoria, neutral]
        write_data_row(ws, row, values, (row - data_start) % 2 == 0)
        row += 1

        totals["total"] += total
        totals["conforme"] += conforme
        totals["melhoria"] += melhoria
        totals["neutral"] += neutral

    # Total row
    total_values = ["TOTAL", totals["total"], totals["conforme"], totals["melhoria"], totals["neutral"]]
    for col_idx, val in enumerate(total_values, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = val
        cell.font = Font(name="Arial", size=10, bold=True, color=OFF_WHITE)
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    auto_width(ws, zone_cols)

    # --- Tabela 2: Desconexoes por Impacto ---
    row += 2
    ws.cell(row=row, column=1, value="Desconexoes por Impacto").font = Font(
        name="Arial", size=12, bold=True, color=VERDE_CAQUI
    )
    row += 1

    impact_cols = ["Impacto", "Quantidade", "Percentual"]
    write_table_header(ws, row, impact_cols)
    row += 1

    total_disc = len(disconnections)
    for impact_label in ["Alto", "Medio", "Baixo"]:
        count = sum(
            1 for d in disconnections
            if (d.get("impact", "").lower().strip() in (impact_label.lower(), impact_label.lower().replace("e", "é")))
        )
        pct = f"{(count / total_disc * 100):.0f}%" if total_disc > 0 else "0%"
        values = [impact_label, count, pct]
        write_data_row(ws, row, values, row % 2 == 0)

        # Formatacao
        impact_cell = ws.cell(row=row, column=1)
        apply_impact_style(impact_cell, impact_label)
        row += 1


def translate_zone(zone):
    """Traduz zona para portugues."""
    translations = {
        "input": "Entrada",
        "output": "Saida",
        "regulation": "Regulacao",
        "support": "Suporte",
    }
    return translations.get(zone, zone)


def translate_status(status):
    """Traduz status para portugues."""
    translations = {
        "conforme": "Conforme",
        "melhoria": "Melhoria",
        "neutral": "Nao Avaliado",
    }
    return translations.get(status, status)


def main():
    parser = argparse.ArgumentParser(description="Gera Tabela de Desconexoes Excel (M7-2026)")
    parser.add_argument("--input", required=True, help="Caminho do JSON de entrada")
    parser.add_argument("--output", required=True, help="Caminho do arquivo .xlsx de saida")
    parser.add_argument("--logo", default=None, help="Caminho do logo M7 (PNG)")
    args = parser.parse_args()

    # Ler JSON
    if args.input == "-":
        data = json.load(sys.stdin)
    else:
        if not os.path.isfile(args.input):
            print(f"ERRO: Arquivo de entrada nao encontrado: {args.input}", file=sys.stderr)
            sys.exit(1)
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)

    # Validar estrutura
    if "metadata" not in data:
        print("ERRO: JSON deve conter 'metadata'", file=sys.stderr)
        sys.exit(1)
    if "interfaces" not in data:
        print("ERRO: JSON deve conter 'interfaces'", file=sys.stderr)
        sys.exit(1)

    # Criar workbook
    wb = Workbook()

    # Logo path
    logo_path = args.logo
    if logo_path and not os.path.isfile(logo_path):
        print(f"AVISO: Logo nao encontrado em {logo_path}, continuando sem logo", file=sys.stderr)
        logo_path = None

    # Criar abas
    create_disconnections_sheet(wb, data, logo_path)
    create_all_interfaces_sheet(wb, data, logo_path)
    create_summary_sheet(wb, data, logo_path)

    # Salvar
    wb.save(args.output)
    print(f"Tabela gerada: {args.output}")


if __name__ == "__main__":
    main()
